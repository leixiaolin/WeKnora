"""
Excel Parser Module

This module provides functionality to parse Excel files (.xlsx, .xls) into
structured Document objects with text content and chunks. It supports multiple
sheets and handles various Excel formats using pandas.
"""
import logging
import re
from io import BytesIO
from typing import List, Tuple

import pandas as pd

from docreader.models.document import Chunk, Document
from docreader.parser.base_parser import BaseParser
from docreader.parser.excel_convert import (
    convert_excel_to_xlsx_bytes,
    detect_excel_format,
    engine_for_format,
    normalize_excel_bytes,
)
from docreader.parser.xlsx_merge import fill_merged_cells_xlsx
from docreader.parser.xlsx_repair import repair_xlsx_bytes

logger = logging.getLogger(__name__)

# Pattern to detect Excel image function strings that should be excluded from
# parsed text content.  WPS uses =DISPIMG("ID",mode) to embed images in cells;
# when opened by other tools the formula may appear as plain text prefixed with
# "_xlfn." or "=".  Office 365 uses =_xlfn.IMAGE(url, ...) similarly.
# The _xlfn. prefix is optional — WPS may omit it (e.g. =DISPIMG("ID",1)).
_IMAGE_FUNC_RE = re.compile(
    r"^=?(_xlfn\.)?(DISPIMG|IMAGE)\(", re.IGNORECASE
)
_HEADER_KEYWORD_RE = re.compile(
    r"(序号|编号|代码|姓名|名称|学院|专业|方向|类别|类型|成绩|分数|备注|"
    r"考生|录取|拟录取|准考证|身份证|电话|邮箱|时间|日期|状态)"
)


def _is_image_function(value: object) -> bool:
    """Return True if *value* looks like an embedded-image function string."""
    if not isinstance(value, str):
        return False
    return _IMAGE_FUNC_RE.match(value) is not None


def _cell_text(value: object) -> str:
    if pd.isna(value) or _is_image_function(value):
        return ""
    return str(value).strip()


def _looks_numeric(value: str) -> bool:
    if value == "":
        return False
    return re.fullmatch(r"[-+]?\d+(?:\.\d+)?", value) is not None


def _infer_header_labels(df: pd.DataFrame) -> Tuple[dict[object, str], int]:
    """Infer semantic column labels while keeping positional A/B/C columns.

    Many XLSX files put a merged title row before the real header row. Pandas
    with header=None preserves all rows, which is good for recall, but row
    chunks then only say "A/B/C" and lose labels like "拟录取专业". This helper
    finds a likely header row and lets later data rows carry both forms.
    """
    best_score = 0
    best_labels: dict[object, str] = {}
    best_row_idx = -1
    max_scan_rows = min(len(df), 20)

    for row_idx in range(max_scan_rows):
        row = df.iloc[row_idx]
        values = [_cell_text(v) for v in row.tolist()]
        non_empty = [v for v in values if v]
        distinct = {v for v in non_empty}
        if len(non_empty) < 2 or len(distinct) < 2:
            continue

        numeric_count = sum(1 for v in non_empty if _looks_numeric(v))
        keyword_count = sum(1 for v in non_empty if _HEADER_KEYWORD_RE.search(v))
        short_text_count = sum(1 for v in non_empty if len(v) <= 24 and not _looks_numeric(v))

        # Header rows are usually compact text labels. Requiring at least one
        # domain-ish keyword avoids treating ordinary first data rows as headers.
        if keyword_count == 0 or short_text_count < 2:
            continue
        if numeric_count > len(non_empty) / 2:
            continue

        score = keyword_count * 3 + short_text_count - numeric_count
        if score <= best_score:
            continue

        labels = {}
        for col, value in zip(df.columns, values):
            if value:
                labels[col] = value
        best_score = score
        best_labels = labels
        best_row_idx = row_idx

    return best_labels, best_row_idx


class ExcelParser(BaseParser):
    """Parser for Excel files (.xlsx, .xls).
    
    This parser extracts text content from Excel files by processing all sheets
    and converting each row into a structured text format. Each row becomes a
    separate chunk with key-value pairs.
    
    Features:
        - Supports multiple sheets in a single Excel file
        - Automatically removes completely empty rows
        - Converts each row to "column: value" format
        - Creates individual chunks for each row for better granularity
        
    Example:
        >>> parser = ExcelParser()
        >>> with open("data.xlsx", "rb") as f:
        ...     content = f.read()
        ...     document = parser.parse_into_text(content)
        >>> print(document.content)
        Name: John,Age: 30,City: NYC
        Name: Jane,Age: 25,City: LA
    """
    
    def parse_into_text(self, content: bytes) -> Document:
        """Parse Excel file bytes into a Document object.
        
        Args:
            content: Raw bytes of the Excel file
            
        Returns:
            Document: Parsed document containing:
                - content: Full text with all rows from all sheets
                - chunks: List of Chunk objects, one per row
                
        Note:
            - Empty rows (all NaN values) are automatically skipped
            - Each row is formatted as: "col1: val1,col2: val2,..."
            - Chunks maintain sequential ordering across all sheets
        """
        chunks: List[Chunk] = []
        text: List[str] = []
        start, end = 0, 0

        excel_file = _open_excel_file(content, file_type=self.file_type)
        
        # Process each sheet in the Excel file
        for excel_sheet_name in excel_file.sheet_names:
            df = _read_sheet_dataframe(excel_file, excel_sheet_name)
            # Remove rows where all values are NaN (completely empty rows)
            df.dropna(how="all", inplace=True)
            header_labels, header_row_idx = _infer_header_labels(df)

            # Process each row in the DataFrame
            for row_pos, (_, row) in enumerate(df.iterrows()):
                page_content = []
                # Build key-value pairs for non-null values
                for k, v in row.items():
                    if pd.notna(v) and not _is_image_function(v):
                        value = str(v)
                        label = header_labels.get(k, "") if row_pos > header_row_idx else ""
                        if label and label != value.strip():
                            page_content.append(f"{k}: {value} [列名: {label}]")
                        else:
                            page_content.append(f"{k}: {value}")
                
                # Skip rows with no valid content
                if not page_content:
                    continue
                
                # Format row as comma-separated key-value pairs
                content_row = ",".join(page_content) + "\n"
                end += len(content_row)
                text.append(content_row)
                
                # Create a chunk for this row with position tracking
                chunks.append(
                    Chunk(content=content_row, seq=len(chunks), start=start, end=end)
                )
                start = end

        # Combine all text and return as Document
        return Document(content="".join(text), chunks=chunks)


def _read_sheet_dataframe(excel_file: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    """Read a worksheet into a DataFrame with stable column labels."""
    from openpyxl.utils import get_column_letter

    # XLSX is preprocessed (merge fill); use A/B/C column letters and keep row 1 as data.
    if excel_file.engine == "openpyxl":
        df = excel_file.parse(sheet_name=sheet_name, header=None)
        df.columns = [get_column_letter(idx + 1) for idx in range(len(df.columns))]
        return df

    df = excel_file.parse(sheet_name=sheet_name, header=0)
    if df.empty:
        df = excel_file.parse(sheet_name=sheet_name, header=None)
        df.columns = [get_column_letter(idx + 1) for idx in range(len(df.columns))]
    elif any(str(col).startswith("Unnamed:") for col in df.columns):
        df = excel_file.parse(sheet_name=sheet_name, header=None)
        df.columns = [get_column_letter(idx + 1) for idx in range(len(df.columns))]
    return df


def _prepare_xlsx_bytes(data: bytes) -> bytes:
    repaired = repair_xlsx_bytes(data)
    if repaired is not None:
        data = repaired
    return fill_merged_cells_xlsx(data)


def _open_excel_file(content: bytes, file_type: str | None = None) -> pd.ExcelFile:
    """Open an Excel workbook with explicit engine selection and fallbacks."""
    data = content
    converted_via_soffice = False

    while True:
        ext = detect_excel_format(data)
        if ext is None:
            if converted_via_soffice:
                raise ValueError(
                    "Excel file format cannot be determined, you must specify an engine manually."
                )
            try:
                data = normalize_excel_bytes(data, file_type=file_type)
            except ValueError as exc:
                raise ValueError(
                    "Excel file format cannot be determined, you must specify an engine manually."
                ) from exc
            converted_via_soffice = True
            continue

        if ext == "ods":
            converted = convert_excel_to_xlsx_bytes(data, suffix=".ods")
            if converted:
                data = converted
                continue

        engine = engine_for_format(ext)
        if ext == "xlsx":
            data = _prepare_xlsx_bytes(data)
            engine = "openpyxl"
        try:
            return pd.ExcelFile(BytesIO(data), engine=engine)
        except ImportError as exc:
            raise ValueError(
                f"Excel engine {engine!r} is not available for .{ext} files"
            ) from exc
        except KeyError as exc:
            if "sharedStrings.xml" not in str(exc) or engine != "openpyxl":
                raise
            repaired = repair_xlsx_bytes(data)
            if repaired is None:
                raise
            logger.info("Repaired XLSX sharedStrings packaging before parse")
            data = _prepare_xlsx_bytes(repaired)
            continue
        except ValueError as exc:
            if converted_via_soffice or "cannot be determined" not in str(exc):
                raise
            try:
                data = normalize_excel_bytes(content, file_type=file_type)
            except ValueError:
                raise
            converted_via_soffice = True
            continue


if __name__ == "__main__":
    # Example usage: Parse an Excel file and display results
    logging.basicConfig(level=logging.DEBUG)

    # Specify the path to your Excel file
    your_file = "/path/to/your/file.xlsx"
    parser = ExcelParser()
    
    # Read and parse the Excel file
    with open(your_file, "rb") as f:
        content = f.read()
        document = parser.parse_into_text(content)
        
        # Display the full document content
        logger.error(document.content)

        # Display the first chunk as an example
        for chunk in document.chunks:
            logger.error(chunk.content)
            break  # Only show the first chunk
